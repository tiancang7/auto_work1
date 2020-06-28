
from openpyxl import load_workbook
from win32com.client import Dispatch
import os
import time

def just_open(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()


end_price = float(input('请输入最终谈价金额：'))
souce_file = 'F:\\谈价会议纪要\\程序化谈价\\'
input_name = input('请输入待谈价文件名：')
data_file = souce_file + input_name


while True:
    just_open(data_file)
    wb = load_workbook(data_file, data_only=True)
    ws = wb.worksheets[0]
    last_tatol = ws.cell(63, 6).value
    last_man_time = ws.cell(10, 6).value
    wb.close()

    if last_tatol>end_price:
        print(last_tatol)
        workbook_ = load_workbook(data_file)
        sheetnames = workbook_.get_sheet_names()  # 获得表单名字
        sheet = workbook_.get_sheet_by_name(sheetnames[0])
        sheet['F10'] = last_man_time - 0.5
        workbook_.save(data_file)
        time.sleep(1)
    else:
        workbook_ = load_workbook(data_file)
        sheetnames = workbook_.get_sheet_names()  # 获得表单名字
        sheet = workbook_.get_sheet_by_name(sheetnames[0])
        sheet['F10'] = last_man_time + 0.5
        sheet['A1'] = '维修项目预算审核'
        sheet['A80'] = '                                     该项目维修造价审核后： ' + str(end_price) + ' 元'
        workbook_.save(data_file)
        break

os.startfile(data_file,'print')
